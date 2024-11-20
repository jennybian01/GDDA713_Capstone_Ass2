from re import X
from turtle import width
from matplotlib import axes, legend_handler
from matplotlib.figure import Figure
from matplotlib.ticker import FuncFormatter
import numpy as np
import sqlite3
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import pandas as pd
import panel as pn


df = pd.read_excel("PMF_Raw Date_QueenSt_PM10.xlsx")

# Replace all missing values with 0
pd.DataFrame.fillna(0, inplace=True)

# Save the modified DataFrame back to an Excel file
pd.DataFrame.to_excel('PMF_Raw Date_QueenSt_PM10.xlsx', index=False)

# create new workbook
wb_new = Workbook()
ws_new = wb_new.active

# set up titles
columns_titles = {
    'A': 'Species',
    'B': 'Soils_C', 'C': 'Soils_A', 'D': 'Soils_error', 'E': 'Soils_DISP', 'F': 'Soils_E',
    'G': 'Sulphate_C', 'H': 'Sulphate_A', 'I': 'Sulphate_error', 'J': 'Sulphate_DISP', 'K': 'Sulphate_E',
    'L': 'Diesel vehicles_C', 'M': 'Diesel vehicles_A', 'N': 'Diesel vehicles_error', 'O': 'Diesel vehicles_DISP', 'P': 'Diesel vehicles_E',
    'Q': 'Marine aerosol_C', 'R': 'Marine aerosol_A', 'S': 'Marine aerosol_error', 'T': 'Marine aerosol_DISP', 'U': 'Marine aerosol_E',
    'V': 'Biomass burning_C', 'W': 'Biomass burning_A', 'X': 'Biomass burning_error', 'Y': 'Biomass burning_DISP', 'Z': 'Biomass burning_E',
    'AA': 'Petrol vehicles_C', 'AB': 'Petrol vehicles_A', 'AC': 'Petrol vehicles_error', 'AD': 'Petrol vehicles_DISP', 'AE': 'Petrol vehicles_E',
    'AF': 'Construction_C', 'AG': 'Construction_A', 'AH': 'Construction_error', 'AI': 'Construction_DISP', 'AJ': 'Construction_E'
}
for col, title in columns_titles.items():
    ws_new[col + '1'] = title

# add exsiting workbook
wb = load_workbook('PMF_Raw Date_QueenSt_PM10.xlsx')
ws = wb.active

# copy raw1.xlsx to raw3.xlsx

for i in range(57, 77):
    value = ws['A' + str(i)].value
    ws_new['A' + str(i - 55)] = value

# save into new workbook
wb_new.save('PMF_Raw Date_QueenSt_PM10_1.xlsx')

# Cache data to improve dashboard performance
if 'data' not in pn.state.cache.keys():
    # Load data from the Excel file
    df = pd.read_excel('PMF_Raw Date_QueenSt_PM10_1.xlsx')
    pn.state.cache['data'] = df.copy()
else:
    df = pn.state.cache['data']

#Soil and road dust
# Load the existing workbook for raw1.xlsx
wb1 = load_workbook('PMF_Raw Date_QueenSt_PM10.xlsx')
ws1 = wb1.active

# Load the existing workbook for raw3.xlsx
wb3 = load_workbook('PMF_Raw Date_QueenSt_PM10_1.xlsx')
ws3 = wb3.active

# Get the divisor value which is in cell B56
divisor = ws1['B56'].value

 # Calculate the values for column B in raw3.xlsx using the values from column B in raw1.xlsx
for i in range(57, 77):
    # Get the value from column B in raw1.xlsx
    value = ws1[f'B{i}'].value
    # Calculate the result to be written in raw3.xlsx
    # Avoid division by zero
    result = (value / divisor) / 1000 if divisor else None
    # Write the result to column B in raw3.xlsx starting from row 2
    ws3[f'B{i-55}'] = result

# Calculate the values for column C in raw3.xlsx using the values from column N in raw1.xlsx
for i in range(57, 77):
    # Get the value from column N in raw1.xlsx
    value_n = ws1[f'N{i}'].value
    # Calculate the result for column C in raw3.xlsx
    result_n = (value_n / divisor) / 1000 if divisor else None
    # Write the result to column C in raw3.xlsx starting from row 2
    ws3[f'C{i-55}'] = result_n

# Calculate the values for column D in raw3.xlsx using the values from column M in raw1.xlsx
for i in range(57, 77):
    # Get the value from column M in raw1.xlsx
    value_m = ws1[f'M{i}'].value
    # Calculate the result for column D in raw3.xlsx
    result_m = (value_m / divisor) / 1000 if divisor else None
    # Write the result to column D in raw3.xlsx starting from row 2
    ws3[f'D{i-55}'] = result_m

# Calculate the values for column E in raw3.xlsx using the values from column O in raw1.xlsx
for i in range(57, 77):
    # Get the value from column O in raw1.xlsx
    value_o = ws1[f'O{i}'].value
    # Calculate the result for column E in raw3.xlsx
    result_o = (value_o / divisor) / 1000 if divisor else None
    # Write the result to column E in raw3.xlsx starting from row 2
    ws3[f'E{i-55}'] = result_o

# Copy values from raw1.xlsx column B (from B225 to B244) to raw3.xlsx column F (from F2 to F21)
for i in range(225, 245):
    # Get the value from column B in raw1.xlsx
    value = ws1[f'B{i}'].value
    # Write the value to column F in raw3.xlsx
    ws3[f'F{i-223}'] = value

# Save the updated raw3.xlsx workbook
wb3.save('PMF_Raw Date_QueenSt_PM10_1.xlsx')

#Sulphate/marine diesel
# Load the existing workbook for raw1.xlsx
wb1 = load_workbook('PMF_Raw Date_QueenSt_PM10.xlsx')
ws1 = wb1.active

# Load the existing workbook for raw3.xlsx
wb3 = load_workbook('PMF_Raw Date_QueenSt_PM10_1.xlsx')
ws3 = wb3.active

# Get the divisor value which is in cell B200
divisor_1 = ws1['B200'].value

# Calculate the values for column G in raw3.xlsx using the values from column B in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column B in raw1.xlsx (B201 to B220)
    value_g = ws1[f'B{200+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_g = (value_g / divisor_1) / 1000 if divisor else None
    # Write the result to column G in raw3.xlsx starting from row 2
    ws3[f'G{i+1}'] = result_g

# Calculate the values for column H in raw3.xlsx using the values from column N in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column N in raw1.xlsx (N201 to N220)
    value_h = ws1[f'N{200+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_h = (value_h / divisor_1) / 1000 if divisor else None
    # Write the result to column H in raw3.xlsx starting from row 2
    ws3[f'H{i+1}'] = result_h

# Calculate the values for column I in raw3.xlsx using the values from column M in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column M in raw1.xlsx (M201 to M220)
    value_I = ws1[f'M{200+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_I = (value_I / divisor_1) / 1000 if divisor else None
    # Write the result to column I in raw3.xlsx starting from row 2
    ws3[f'I{i+1}'] = result_I

# Calculate the values for column J in raw3.xlsx using the values from column O in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column O in raw1.xlsx (O201 to O220)
    value_j = ws1[f'O{200+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_j = (value_j / divisor_1) / 1000 if divisor else None
    # Write the result to column J in raw3.xlsx starting from row 2
    ws3[f'j{i+1}'] = result_j

# Copy values from raw1.xlsx column B (from B369 to B388) to raw3.xlsx column F (from K2 to K21)
for i in range(369, 389):
    # Get the value from column B in raw1.xlsx
    value = ws1[f'B{i}'].value
    # Write the value to column K in raw3.xlsx
    ws3[f'K{i-367}'] = value

    # Save the updated raw3.xlsx workbook
    wb3.save('PMF_Raw Date_QueenSt_PM10_1.xlsx')

    # Diesel vehicles

    # Load the existing workbook for raw1.xlsx
    wb1 = load_workbook('PMF_Raw Date_QueenSt_PM10.xlsx')
    ws1 = wb1.active

    # Load the existing workbook for raw3.xlsx
    wb3 = load_workbook('PMF_Raw Date_QueenSt_PM10_1.xlsx')
    ws3 = wb3.active

    # Get the divisor value which is in cell B200
    divisor_3 = ws1['B104'].value

# Calculate the values for column L in raw3.xlsx using the values from column B in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column B in raw1.xlsx (B105 to B124)
    value_l = ws1[f'B{104+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_l = (value_l / divisor_3) / 1000 if divisor else None
    # Write the result to column L in raw3.xlsx starting from row 2
    ws3[f'l{i+1}'] = result_l

# Calculate the values for column M in raw3.xlsx using the values from column N in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column N in raw1.xlsx (N105 to N124)
    value_m = ws1[f'N{104+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_m = (value_m / divisor_3) / 1000 if divisor else None
    # Write the result to column H in raw3.xlsx starting from row 2
    ws3[f'm{i+1}'] = result_m

# Calculate the values for column N in raw3.xlsx using the values from column M in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column M in raw1.xlsx (M105 to M124)
    value_n = ws1[f'M{104+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_n = (value_n / divisor_3) / 1000 if divisor else None
    # Write the result to column I in raw3.xlsx starting from row 2
    ws3[f'n{i+1}'] = result_n

# Calculate the values for column O in raw3.xlsx using the values from column O in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column O in raw1.xlsx (O105 to O124)
    value_o = ws1[f'O{104+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_o = (value_o / divisor_3) / 1000 if divisor else None
    # Write the result to column J in raw3.xlsx starting from row 2
    ws3[f'o{i+1}'] = result_o

# Copy values from raw1.xlsx column B (from B369 to B388) to raw3.xlsx column F (from K2 to K21)
for i in range(273, 293):
    # Get the value from column B in raw1.xlsx
    value = ws1[f'B{i}'].value
    # Write the value to column K in raw3.xlsx
    ws3[f'P{i-271}'] = value

# Save the updated raw3.xlsx workbook
wb3.save('PMF_Raw Date_QueenSt_PM10_1.xlsx')

#Sea salt
# Load the existing workbook for raw1.xlsx
wb1 = load_workbook('PMF_Raw Date_QueenSt_PM10.xlsx')
ws1 = wb1.active

# Load the existing workbook for raw3.xlsx
wb3 = load_workbook('PMF_Raw Date_QueenSt_PM10_1.xlsx')
ws3 = wb3.active

# Get the divisor value which is in cell B80
divisor_4 = ws1['B80'].value

# Calculate the values for column Q in raw3.xlsx using the values from column B in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column B in raw1.xlsx (B81 to B100)
    value_q = ws1[f'B{80+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_q = (value_q / divisor_4) / 1000 if divisor else None
    # Write the result to column L in raw3.xlsx starting from row 2
    ws3[f'q{i+1}'] = result_q

# Calculate the values for column R in raw3.xlsx using the values from column N in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column N in raw1.xlsx (N81 to N100)
    value_r = ws1[f'N{80+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_r = (value_r / divisor_4) / 1000 if divisor else None
    # Write the result to column H in raw3.xlsx starting from row 2
    ws3[f'r{i+1}'] = result_r

# Calculate the values for column S in raw3.xlsx using the values from column M in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column M in raw1.xlsx (M81 to M100)
    value_s = ws1[f'M{80+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_s = (value_s / divisor_4) / 1000 if divisor else None
    # Write the result to column I in raw3.xlsx starting from row 2
    ws3[f's{i+1}'] = result_s

# Calculate the values for column T in raw3.xlsx using the values from column O in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column O in raw1.xlsx (O81 to O100)
    value_t = ws1[f'O{80+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_t = (value_t / divisor_4) / 1000 if divisor else None
    # Write the result to column J in raw3.xlsx starting from row 2
    ws3[f't{i+1}'] = result_t

# Copy values from raw1.xlsx column B (from B249 to B269) to raw3.xlsx column U
for i in range(249, 269):
    # Get the value from column B in raw1.xlsx
    value = ws1[f'B{i}'].value
    # Write the value to column U in raw3.xlsx
    ws3[f'U{i-247}'] = value

# Save the updated raw3.xlsx workbook
wb3.save('PMF_Raw Date_QueenSt_PM10_1.xlsx')

#Biomass burning

# Load the existing workbook for raw1.xlsx
wb1 = load_workbook('PMF_Raw Date_QueenSt_PM10.xlsx')
ws1 = wb1.active

# Load the existing workbook for raw3.xlsx
wb3 = load_workbook('PMF_Raw Date_QueenSt_PM10_1.xlsx')
ws3 = wb3.active

# Get the divisor value which is in cell B80
divisor_4 = ws1['B128'].value

# Calculate the values for column V in raw3.xlsx using the values from column B in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column B in raw1.xlsx (B129 to B148)
    value_v = ws1[f'B{128+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_v = (value_v / divisor_4) / 1000 if divisor_4 else None
    # Write the result to column V in raw3.xlsx starting from row 2
    ws3[f'V{i+1}'] = result_v

# Calculate the values for column W in raw3.xlsx using the values from column N in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column N in raw1.xlsx (N129 to N148)
    value_w = ws1[f'N{128+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_w = (value_w / divisor_4) / 1000 if divisor_4 else None
    # Write the result to column W in raw3.xlsx starting from row 2
    ws3[f'W{i+1}'] = result_w

# Calculate the values for column X in raw3.xlsx using the values from column M in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column M in raw1.xlsx (M129 to M148)
    value_x = ws1[f'M{128+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_x = (value_x / divisor_4) / 1000 if divisor_4 else None
    # Write the result to column X in raw3.xlsx starting from row 2
    ws3[f'X{i+1}'] = result_x

# Calculate the values for column Y in raw3.xlsx using the values from column O in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column O in raw1.xlsx (O129 to O148)
    value_y = ws1[f'O{128+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_y = (value_y / divisor_4) / 1000 if divisor_4 else None
    # Write the result to column Y in raw3.xlsx starting from row 2
    ws3[f'Y{i+1}'] = result_y

# Copy values from raw1.xlsx column B (from B369 to B388) to raw3.xlsx column Z
for i in range(297, 317):
    # Get the value from column B in raw1.xlsx
    value = ws1[f'B{i}'].value
    # Write the value to column Z in raw3.xlsx
    ws3[f'Z{i-295}'] = value

# Save the updated raw3.xlsx workbook
wb3.save('PMF_Raw Date_QueenSt_PM10_1.xlsx')

# Petrol vehicles

# Load the existing workbook for raw1.xlsx
wb1 = load_workbook('PMF_Raw Date_QueenSt_PM10.xlsx')
ws1 = wb1.active

# Load the existing workbook for raw3.xlsx
wb3 = load_workbook('PMF_Raw Date_QueenSt_PM10_1.xlsx')
ws3 = wb3.active

# Get the divisor value which is in cell B80
divisor_4 = ws1['B176'].value

# Calculate the values for column V in raw3.xlsx using the values from column B in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column B in raw1.xlsx (B129 to B148)
    value_aa = ws1[f'B{176+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_aa = (value_aa / divisor_4) / 1000 if divisor_4 else None
    # Write the result to column V in raw3.xlsx starting from row 2
    ws3[f'AA{i+1}'] = result_aa

# Calculate the values for column W in raw3.xlsx using the values from column N in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column N in raw1.xlsx (N129 to N148)
    value_ab = ws1[f'N{176+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_ab = (value_ab / divisor_4) / 1000 if divisor_4 else None
    # Write the result to column W in raw3.xlsx starting from row 2
    ws3[f'AB{i+1}'] = result_ab

# Calculate the values for column X in raw3.xlsx using the values from column M in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column M in raw1.xlsx (M129 to M148)
    value_ac = ws1[f'M{176+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_ac = (value_ac / divisor_4) / 1000 if divisor_4 else None
    # Write the result to column X in raw3.xlsx starting from row 2
    ws3[f'AC{i+1}'] = result_ac

# Calculate the values for column Y in raw3.xlsx using the values from column O in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column O in raw1.xlsx (O129 to O148)
    value_ad = ws1[f'O{176+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_ad = (value_ad / divisor_4) / 1000 if divisor_4 else None
    # Write the result to column Y in raw3.xlsx starting from row 2
    ws3[f'AD{i+1}'] = result_ad

# Copy values from raw1.xlsx column B (from B369 to B388) to raw3.xlsx column Z
for i in range(345, 365):
    # Get the value from column B in raw1.xlsx
    value = ws1[f'B{i}'].value
    # Write the value to column Z in raw3.xlsx
    ws3[f'AE{i-343}'] = value

# Save the updated raw3.xlsx workbook
wb3.save('PMF_Raw Date_QueenSt_PM10_1.xlsx')

# Construction
# Load the existing workbook for raw1.xlsx
wb1 = load_workbook('PMF_Raw Date_QueenSt_PM10.xlsx')
ws1 = wb1.active

# Load the existing workbook for raw3.xlsx
wb3 = load_workbook('PMF_Raw Date_QueenSt_PM10_1.xlsx')
ws3 = wb3.active

# Get the divisor value which is in cell B80
divisor_4 = ws1['B152'].value

# Calculate the values for column V in raw3.xlsx using the values from column B in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column B in raw1.xlsx (B129 to B148)
    value_af = ws1[f'B{152+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_af = (value_af / divisor_4) / 1000 if divisor_4 else None
    # Write the result to column V in raw3.xlsx starting from row 2
    ws3[f'AF{i+1}'] = result_af

# Calculate the values for column W in raw3.xlsx using the values from column N in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column N in raw1.xlsx (N129 to N148)
    value_ag = ws1[f'N{152+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_ag = (value_ag / divisor_4) / 1000 if divisor_4 else None
    # Write the result to column W in raw3.xlsx starting from row 2
    ws3[f'AG{i+1}'] = result_ag

# Calculate the values for column X in raw3.xlsx using the values from column M in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column M in raw1.xlsx (M129 to M148)
    value_ah = ws1[f'M{152+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_ah = (value_ah / divisor_4) / 1000 if divisor_4 else None
    # Write the result to column X in raw3.xlsx starting from row 2
    ws3[f'AH{i+1}'] = result_ah

# Calculate the values for column Y in raw3.xlsx using the values from column O in raw1.xlsx
for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in raw3.xlsx
    # Get the value from column O in raw1.xlsx (O129 to O148)
    value_ai = ws1[f'O{152+i}'].value
    # Calculate the result to be written in raw3.xlsx
    result_ai = (value_ai / divisor_4) / 1000 if divisor_4 else None
    # Write the result to column Y in raw3.xlsx starting from row 2
    ws3[f'AI{i+1}'] = result_ai

# Copy values from raw1.xlsx column B (from B369 to B388) to raw3.xlsx column Z
for i in range(321, 341):
    # Get the value from column B in raw1.xlsx
    value = ws1[f'B{i}'].value
    # Write the value to column Z in raw3.xlsx
    ws3[f'AJ{i-319}'] = value

# Save the updated raw3.xlsx workbook
wb3.save('PMF_Raw Date_QueenSt_PM10_1.xlsx')

# Cache data to improve dashboard performance
if 'data' not in pn.state.cache.keys():
    # Load data from the Excel file
    df = pd.read_excel('PMF_Raw Date_QueenSt_PM10_1.xlsx')
    pn.state.cache['data'] = df.copy()
else:
    df = pn.state.cache['data']

#PMF_Raw Date_QueenSt_PM10_1.xslsx Sheet2 for Pie Chart
# Load the existing workbook for PMF_Raw Date_QueenSt_PM10.xlsx
wb1 = load_workbook('PMF_Raw Date_QueenSt_PM10.xlsx')
ws1 = wb1.active

# Load the existing workbook for PMF_Raw Date_QueenSt_PM10_1.xlsx
wb3 = load_workbook('PMF_Raw Date_QueenSt_PM10_1.xlsx')

# Create a new worksheet in the existing workbook
ws3 = wb3.create_sheet(title='Sheet2')

# Set the values for A1 and B1
ws3.cell(row=1, column=1, value='Sources')  # Set 'Sources' in A1
ws3.cell(row=1, column=2, value='Percentage')  # Set 'Percentage' in B1

# Extract the required data from ws1
data_to_extract = {
    'B2': ws1.cell(row=224, column=2).value,  # B224
    'B3': ws1.cell(row=248, column=2).value,  # B248
    'B4': ws1.cell(row=272, column=2).value,  # B272
    'B5': ws1.cell(row=296, column=2).value,  # B296
    'B6': ws1.cell(row=320, column=2).value,  # B320
    'B7': ws1.cell(row=344, column=2).value,  # B344
    'B8': ws1.cell(row=368, column=2).value   # B368
}

# Data labels for column A
labels = ['Soil/Road dust',
    'Sea salt',
    'Diesel vehicles',
    'Biomass burning',
    'Construction',
    'Petrol vehicles',
    'Sulphate/Marine diesel'
]

# Update ws3 with the extracted data
for i, (key, value) in enumerate(data_to_extract.items()):
    ws3.cell(row=i + 2, column=2, value=value)  # Write to column B (2)
    ws3.cell(row=i + 2, column=1, value=labels[i])  # Write to column A (1)

# Save the updated workbook
wb3.save('PMF_Raw Date_QueenSt_PM10_1.xlsx')

#data preparation for species
sources = {
    'Soils': ['C', 'A', 'error', 'DISP', 'E'],
    'Sulphate': ['C', 'A', 'error', 'DISP', 'E'],
    'Diesel vehicles': ['C', 'A', 'error', 'DISP', 'E'],
    'Marine aerosol': ['C', 'A', 'error', 'DISP', 'E'],
    'Biomass burning': ['C', 'A', 'error', 'DISP', 'E'],
    'Petrol vehicles': ['C', 'A', 'error', 'DISP', 'E'],
    'Construction': ['C', 'A', 'error', 'DISP', 'E']
}
#create a dictionary for splitting df
dfs = {}
for source, measurements in sources.items():
    # conduct a list of pollution sources
    columns = ['Species'] + [f"{source}_{m}" for m in measurements]
    # rename each df name
    dfs[source] = df[columns].copy()
    #rename each column name
    new_columns = ['Species', "Concentration", "Average", "Error", "Dispersion", "Exceedance"]
    dfs[source].columns = new_columns

    df1 = dfs['Soils']
    df2 = dfs['Sulphate']
    df3 = dfs['Diesel vehicles']
    df4 = dfs['Marine aerosol']
    df5 = dfs['Biomass burning']
    df6 = dfs['Petrol vehicles']
    df7 = dfs['Construction']

#Create SQLite DB for species
def insert_species(df, conn):
    cursor = conn.cursor()

    # insert species to species table
    df[['Species']].to_sql('Species', conn, if_exists='append', index=False)

    # confirm
    conn.commit()

    # connect to SQLite database and insert to species table
    conn = sqlite3.connect('environmental_data.db')

    insert_species(df, conn)
    conn.close()

    #Measurement Table

    # connect to SQLite database
    conn = sqlite3.connect('environmental_data.db')
    cursor = conn.cursor()

    # obtain all SpeciesID
    cursor.execute('SELECT SpeciesID, Species FROM Species')
    species_di = {name: id for id, name in cursor.fetchall()}

# define insert data function
def insert_measurement(df, source_id, conn):
    try:
        cursor = conn.cursor()
        # insert data
        for index, row in df.iterrows():
            # obtain SpeciesID
            species_id = species_id.get(row['Species'])
            if species_id is None:
                print(f"Species {row['Species']} not found in Species table.")
                continue

            # insert Measurement data
            cursor.execute('''
            INSERT INTO Measurement (SampleID, SpeciesID, SourceID, Concentration, Average, Error, Dispersion, Exceedance)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ''', (1, species_id, source_id,
                row['Concentration'], row['Average'], row['Error'], row['Dispersion'], row['Exceedance']))

        # confirm
        conn.commit()
    except sqlite3.Error as e:
        print("SQLite error: ", e)
    except Exception as e:
        print("General error: ", e)

# insert data
source_ids = {
    'Soils': 'P1',
    'Sulphate': 'P2',
    'Diesel vehicles': 'P3',
    'Marine aerosol': 'P4',
    'Biomass burning': 'P5',
    'Petrol vehicles': 'P6',
    'Construction': 'P7'
}

dfs = {
    'Soils': df1,
    'Sulphate': df2,
    'Diesel vehicles': df3,
    'Marine aerosol': df4,
    'Biomass burning': df5,
    'Petrol vehicles': df6,
    'Construction': df7
}

for source, df in dfs.items():
    source_id = source_ids[source]
    insert_measurement(df, source_id, conn)

    # close connection
    conn.close()

    # Extract Data and Connect to SQLite- Pie Chart
    import sqlite3
    from openpyxl import load_workbook

    # Load the existing workbook PMF_Raw Date_QueenSt_PM10_1.xlsx and the specific sheet
    wb = load_workbook('PMF_Raw Date_QueenSt_PM10_1.xlsx')
    ws = wb['Sheet2']  # Make sure you’re pointing to the correct sheet

    # Set up a new SQLite database
    conn = sqlite3.connect('PMF_QueenSt_PM10_1.db')
    cursor = conn.cursor()

    # Create the new table for the data, only selecting column headers without inserting actual values
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS PollutionSources (
            Source TEXT,
            Percentage REAL
        )
    ''')
    conn.commit()

    # Insert Data (Without Actual Values for Percentage)

# Start from the second row, as the first row contains headers
for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
    source = row[0]
    # Insert source with a NULL value for Percentage
    cursor.execute("INSERT INTO PollutionSources (Source, Percentage) VALUES (?, NULL)", (source,))

    conn.commit()
    conn.close()

# Insert Actual Values for Percentage
# Data labels for column A
labels = [
    'Soil/Road dust',
    'Sea salt',
    'Diesel vehicles',
    'Biomass burning',
    'Construction',
    'Petrol vehicles',
    'Sulphate/Marine diesel'
]

# Connect to the SQLite database (assuming it already exists)
conn = sqlite3.connect('PMF_QueenSt_PM10_1.db')
cursor = conn.cursor()

# Insert the extracted data into the SQLite table
for i, label in enumerate(labels):
    percentage = data_to_extract[f'B{i+2}']  # B2 corresponds to index 0, etc.
    cursor.execute("UPDATE PollutionSources SET Percentage = ? WHERE Source = ?", (percentage, label))

    # Commit the changes and close the connection
    conn.commit()
    conn.close()

# Connect to SQLite database
conn = sqlite3.connect('environmental_data.db')
cursor = conn.cursor()

# Species Table Data
species_ids = list(range(1, 21))
species_names = {
    1: 'H', 2: 'BC', 3: 'Na', 4: 'Mg',
    5: 'Al', 6: 'Si', 7: 'S', 8: 'Cl',
    9: 'K', 10: 'Ca', 11: 'Ti', 12: 'V',
    13: 'Mn', 14: 'Fe', 15: 'Co', 16: 'Ni',
    17: 'Cu', 18: 'Zn', 19: 'As', 20: 'Ba'
}

# Query data for each source P1 through P7
sources = ['P1', 'P2', 'P3', 'P4', 'P5', 'P6', 'P7']
source_titles = [
    'Soil/Road Dust', 'Sulphate/Marine Diesel', 'Diesel Vehicles',
    'Marine Aerosol', 'Biomass Burning', 'Petrol Vehicles', 'Construction'
]
data = {}

for source in sources:
    cursor.execute(f"""
    SELECT SpeciesID, Concentration, Average, Error, Dispersion, Exceedance
    FROM Measurement
    WHERE SourceID = '{source}'
    """)
    data[source] = cursor.fetchall()

    # Close the database connection
    conn.close()

def run_visualization():
    # Plot setup
    fig, axes = plt.subplots(len(sources), 1, figsize=(8, 10), sharex=False)
    x = np.arange(len(species_ids))  # Label locations
    width = 0.75  # Width of the bars

    # To keep track of whether to add the legend
    first_plot = True

    # Create handles for the legend to ensure only one set of legend items
    legend_handles = []

for i, (source, title) in enumerate(zip(sources, source_titles)):
    ax = axes[i]
    source_data = data[source]

    # Extract data for plotting
    concentration = [row[1] for row in source_data]
    average = [row[2] for row in source_data]
    error = [row[3] for row in source_data]
    dispersion = [row[4] for row in source_data]
    exceedance = [row[5] for row in source_data]

    # Logarithmic y-axis for concentration
    ax.set_yscale('log')
    ax.set_ylim([1e-4, 1])

    # Format left y-axis ticks to display as scientific notation (10^0, 10^-2, etc.)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda y, _: f'$10^{{{int(np.log10(y))}}}$'))

    # Right y-axis (percentage scale) for Exceedance
    ax2 = ax.twinx()
    ax2.set_ylim([0, 100])
    ax2.set_yticks(np.arange(0, 101, 20))

    # Background vertical lines
for j in value_x:
    ax.vlines(j, 1e-4, 1, colors='gray', linestyles='dashed', lw=0.5)

    # Concentration bars
    bars = ax.bar(value_x, concentration, width, color='lightblue', edgecolor='black')

    # Error and Dispersion as vertical lines
    line = []

for j in range(len(value_x)):
    line.append(ax.plot([value_x[j], value_x[j]], [error[j], dispersion[j]], color='black', lw=1))

    # Average as hollow dots
    avg_line = ax.plot(value_x, average, color='red', marker='o', markersize=6, linestyle='', markerfacecolor='none')

    # Exceedance as green dots
    exceedance_line = ax2.plot(value_x, exceedance, color='black', marker='o', markersize=6, linestyle='')

    # Add labels to the legend only for the first plot (for each source)
    if first_plot:
        # Create handles for the legend
        legend_handler.append(bars[0])  # The light blue bars (Concentration)
        legend_handler.append(avg_line[0])  # The red hollow dots (Average)
        legend_handler.append(exceedance_line[0])  # The black dots (Exceedance)
        legend_handler.append(line[0][0])  # The black lines (Maximum and minimum DISP values)

        first_plot = False

        # Set title for each subplot (top-right corner)
        ax.text(1, 0.8, title, transform=ax.transAxes, ha='right', va='bottom', fontsize=10, weight='bold')

        # Set the x-axis labels
        ax.set_xticks(value_x)
        ax.set_xticklabels([species_names[sid] for sid in species_ids], rotation=0)

        # Add the legend at the bottom of the entire figure (only once)
        Figure.legend(legend_handler, ['Concentration', 'Average', 'Exceedance', 'Maximum and Minimum DISP values'], loc='upper center', bbox_to_anchor=(0.5, -0.0), ncol=4)

        # Adjust the layout to make space for the legend
        plt.tight_layout(pad=0.8)
        plt.show()


        # Connect to the SQLite database
        conn = sqlite3.connect('PMF_QueenSt_PM10_1.db')
        cursor = conn.cursor()

        # Retrieve the data for the pie chart
        cursor.execute("SELECT Source, Percentage FROM PollutionSources")
        data = cursor.fetchall()

        # Close the database connection
        conn.close()

        # Separate the data into labels and sizes
        labels = [row[0] for row in data]
        sizes = [row[1] for row in data]

        # Create a 2D pie chart
        plt.figure(figsize=(10, 7),facecolor=(1, 1, 1, 0.5))

        # Create the pie chart with a slight explode effect for better visibility
        explode = [0.1] * len(labels)
        plt.pie(sizes, explode=explode, labels=labels, autopct='%1.1f%%', startangle=140, pctdistance=0.85)

        # Equal aspect ratio ensures the pie is drawn as a circle
        plt.axis('equal')

        # Add a legend
        plt.legend(labels, loc='upper right', bbox_to_anchor=(1.2, 1))

        # Title for the pie chart
        plt.title('Pollution Sources Breakdown', fontsize=16, pad=24)

        # Show the plot
        plt.show()

        pass