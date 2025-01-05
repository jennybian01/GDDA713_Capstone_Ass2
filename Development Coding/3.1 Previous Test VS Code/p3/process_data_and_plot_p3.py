from matplotlib.ticker import FuncFormatter
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import sqlite3
import os
import panel as pn
from openpyxl import Workbook, load_workbook
from shiny import App
import matplotlib.pyplot as plt
 

def clean_excel_data_p3(input_file, output_file):
    try:
        # Load and clean the Excel data
        df = pd.read_excel(input_file)
        df.fillna(0, inplace=True)
        df.to_excel(output_file, index=False)
        print(f"Data cleaned and saved successfully to {output_file}")
        return df
    except Exception as e:
        print(f"An error occurred while cleaning and saving the data: {e}")
        return None
    
def create_output_file_new(output_file_new):
    try:
        # create new workbook
        wb_new = Workbook()
        ws_new = wb_new.active  

        # set up titles
        columns_titles = {
            'A': 'Species',
            'B': 'Soils_C', 'C': 'Soils_A', 'D': 'Soils_error', 'E': 'Soils_DISP', 'F': 'Soils_E',
            'G': 'Sulphate_C', 'H': 'Sulphate_A', 'I': 'Sulphate_error', 'J': 'Sulphate_DISP', 'K': 'Sulphate_E',
            'L': 'Diesel vehicles_C', 'M': 'Diesel vehicles_A', 'N': 'Diesel vehicles_error', 'O': 'Diesel vehicles_DISP', 'P': 'Diesel vehicles_E',
            'Q': 'Marine aerosol_C', 'R': 'Marine aerosol_A', 'S': 'Marine aerosol_error', 'T': 'Marine aerosol_DISP', 'U': 'Marine aerosol_E',
            'V': 'Biomass burning_C', 'W': 'Biomass burning_A', 'X': 'Biomass burning_error', 'Y': 'Biomass burning_DISP', 'Z': 'Biomass burning_E',
            'AA': 'Petrol vehicles_C', 'AB': 'Petrol vehicles_A', 'AC': 'Petrol vehicles_error', 'AD': 'Petrol vehicles_DISP', 'AE': 'Petrol vehicles_E',
            'AF': 'Construction_C', 'AG': 'Construction_A', 'AH': 'Construction_error', 'AI': 'Construction_DISP', 'AJ': 'Construction_E'
        }
        for col, title in columns_titles.items():
            ws_new[col + '1'] = title  # set up title

        wb_new.save(output_file_new)  # save workbook
        print(f"New workbook created and saved successfully to {output_file_new}")
        return wb_new, ws_new
    except Exception as e:
        print(f"An error occurred while creating the new workbook: {e}")
        return None, None
    
def copy_speciation_cloumn(output_file, output_file_new, wb_new, ws_new):

        try:
            # Load the existing workbook
            wb = load_workbook(output_file)
            ws = wb.active

            # Copy data from ws to ws_new
            for i in range(57, 77):
                value = ws[f'A{i}'].value
                ws_new[f'A{i - 55}'] = value

            # Save into new workbook
            wb_new.save(output_file_new)
            print(f"Data copied successfully from {output_file} to {output_file_new}")
            return wb_new, ws_new
        except Exception as e:
            print(f"An error occurred while copying data: {e}")
            return None, None

def process_data(output_file, output_file_new, wb_new, ws_new):
    try:
        #load the existing workbook
        wb = load_workbook(output_file)
        ws = wb.active


        # Soil Road Dust data (start_row 57, end_row 76, divisor_cell B56)
        divisor = ws['B56'].value

        for i in range(57, 77):  # Loop through rows 57 to 76
            value_b = ws[f'B{i}'].value
            value_n = ws[f'N{i}'].value
            value_m = ws[f'M{i}'].value
            value_o = ws[f'O{i}'].value

            result_b = (value_b / divisor) / 1000 if divisor else None
            result_n = (value_n / divisor) / 1000 if divisor else None
            result_m = (value_m / divisor) / 1000 if divisor else None
            result_o = (value_o / divisor) / 1000 if divisor else None

            ws_new[f'B{i-56}'] = result_b
            ws_new[f'C{i-56}'] = result_n
            ws_new[f'D{i-56}'] = result_m
            ws_new[f'E{i-56}'] = result_o

        # Copy values from output_file (B225 to B244) to new workbook (F2 to F21)
        for i in range(225, 245):
            value = ws[f'B{i}'].value
            ws_new[f'F{i-223}'] = value
################
        # Sulphate Marine Diesel data (start_row 201, end_row 220, divisor_cell B200)
        divisor_2 = ws['B200'].value

        for i in range(1, 21):  # Loop through rows 1 to 20 for Sulphate Marine Diesel
            value_g = ws[f'B{200 + i}'].value
            value_h = ws[f'N{200 + i}'].value
            value_i = ws[f'M{200 + i}'].value
            value_j = ws[f'O{200 + i}'].value

            result_g = (value_g / divisor_2) / 1000 if divisor_2 else None
            result_h = (value_h / divisor_2) / 1000 if divisor_2 else None
            result_i = (value_i / divisor_2) / 1000 if divisor_2 else None
            result_j = (value_j / divisor_2) / 1000 if divisor_2 else None

            ws_new[f'G{i+1}'] = result_g
            ws_new[f'H{i+1}'] = result_h
            ws_new[f'I{i+1}'] = result_i
            ws_new[f'J{i+1}'] = result_j

        # Copy values from output_file (B369 to B388) to new workbook (K2 to K21)
        for i in range(369, 389):
            value = ws[f'B{i}'].value
            ws_new[f'K{i-367}'] = value

#######################            
        # Diesel_vehicles Get the divisor value which is in cell B104
        divisor_3 = ws['B104'].value

        # Calculate the values for columns L, M, N, and O in output_file_new using the values from output_file
        for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in output_file_new
            # Get the values from columns B, N, M, and O in output_file (B105 to B124, N105 to N124, M105 to M124, O105 to O124)
            value_l = ws[f'B{104+i}'].value
            value_m = ws[f'N{104+i}'].value
            value_n = ws[f'M{104+i}'].value
            value_o = ws[f'O{104+i}'].value

            # Calculate the results to be written in output_file_new
            result_l = (value_l / divisor_3) / 1000 if divisor_3 else None
            result_m = (value_m / divisor_3) / 1000 if divisor_3 else None
            result_n = (value_n / divisor_3) / 1000 if divisor_3 else None
            result_o = (value_o / divisor_3) / 1000 if divisor_3 else None

            # Write the results to columns L, M, N, and O in output_file_new starting from row 2
            ws_new[f'L{i+1}'] = result_l
            ws_new[f'M{i+1}'] = result_m
            ws_new[f'N{i+1}'] = result_n
            ws_new[f'O{i+1}'] = result_o

        # Copy values from output_file column B (from B273 to B292) to output_file_new column P (from P2 to P21)
        for i in range(273, 293):
            # Get the value from column B in output_file
            value = ws[f'B{i}'].value
            # Write the value to column P in output_file_new
            ws_new[f'P{i-271}'] = value
######################
        # Sea Salt Get the divisor value which is in cell B80
        divisor_4 = ws['B80'].value

        # Calculate the values for columns Q, R, S, and T in output_file_new using the values from output_file
        for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in output_file_new
            # Get the values from columns B, N, M, and O in output_file (B81 to B100, N81 to N100, M81 to M100, O81 to O100)
            value_q = ws[f'B{80+i}'].value
            value_r = ws[f'N{80+i}'].value
            value_s = ws[f'M{80+i}'].value
            value_t = ws[f'O{80+i}'].value

            # Calculate the results to be written in output_file_new
            result_q = (value_q / divisor_4) / 1000 if divisor_4 else None
            result_r = (value_r / divisor_4) / 1000 if divisor_4 else None
            result_s = (value_s / divisor_4) / 1000 if divisor_4 else None
            result_t = (value_t / divisor_4) / 1000 if divisor_4 else None

            # Write the results to columns Q, R, S, and T in output_file_new starting from row 2
            ws_new[f'Q{i+1}'] = result_q
            ws_new[f'R{i+1}'] = result_r
            ws_new[f'S{i+1}'] = result_s
            ws_new[f'T{i+1}'] = result_t

        # Copy values from output_file column B (from B249 to B269) to output_file_new column U (from U2 to U21)
        for i in range(249, 269):
            # Get the value from column B in output_file
            value = ws[f'B{i}'].value
            # Write the value to column U in output_file_new
            ws_new[f'U{i-247}'] = value
##########################
        # Biomass_burning Get the divisor value which is in cell B128
        divisor_5 = ws['B128'].value

        # Calculate the values for columns V, W, X, and Y in output_file_new using the values from output_file
        for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in output_file_new
            # Get the values from columns B, N, M, and O in output_file (B129 to B148, N129 to N148, M129 to M148, O129 to O148)
            value_v = ws[f'B{128+i}'].value
            value_w = ws[f'N{128+i}'].value
            value_x = ws[f'M{128+i}'].value
            value_y = ws[f'O{128+i}'].value

            # Calculate the results to be written in output_file_new
            result_v = (value_v / divisor_5) / 1000 if divisor_5 else None
            result_w = (value_w / divisor_5) / 1000 if divisor_5 else None
            result_x = (value_x / divisor_5) / 1000 if divisor_5 else None
            result_y = (value_y / divisor_5) / 1000 if divisor_5 else None

            # Write the results to columns V, W, X, and Y in output_file_new starting from row 2
            ws_new[f'V{i+1}'] = result_v
            ws_new[f'W{i+1}'] = result_w
            ws_new[f'X{i+1}'] = result_x
            ws_new[f'Y{i+1}'] = result_y

        # Copy values from output_file column B (from B297 to B316) to output_file_new column Z (from Z2 to Z20)
        for i in range(297, 317):
            # Get the value from column B in output_file
            value = ws[f'B{i}'].value
            # Write the value to column Z in output_file_new
            ws_new[f'Z{i-295}'] = value
#################################
        # Petrol vehicle Get the divisor value which is in cell B176
        divisor_6 = ws['B176'].value

        # Calculate the values for columns AA, AB, AC, and AD in output_file_new using the values from output_file
        for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in output_file_new
            # Get the values from columns B, N, M, and O in output_file (B177 to B196, N177 to N196, M177 to M196, O177 to O196)
            value_aa = ws[f'B{176+i}'].value
            value_ab = ws[f'N{176+i}'].value
            value_ac = ws[f'M{176+i}'].value
            value_ad = ws[f'O{176+i}'].value

            # Calculate the results to be written in output_file_new
            result_aa = (value_aa / divisor_6) / 1000 if divisor_6 else None
            result_ab = (value_ab / divisor_6) / 1000 if divisor_6 else None
            result_ac = (value_ac / divisor_6) / 1000 if divisor_6 else None
            result_ad = (value_ad / divisor_6) / 1000 if divisor_6 else None

            # Write the results to columns AA, AB, AC, and AD in output_file_new starting from row 2
            ws_new[f'AA{i+1}'] = result_aa
            ws_new[f'AB{i+1}'] = result_ab
            ws_new[f'AC{i+1}'] = result_ac
            ws_new[f'AD{i+1}'] = result_ad

        # Copy values from output_file column B (from B345 to B364) to output_file_new column AE (from AE2 to AE21)
        for i in range(345, 365):
            # Get the value from column B in output_file
            value = ws[f'B{i}'].value
            # Write the value to column AE in output_file_new
            ws_new[f'AE{i-343}'] = value
###############################
        # Construction Get the divisor value which is in cell B152
        divisor_7 = ws['B152'].value

        # Calculate the values for columns AF, AG, AH, and AI in output_file_new using the values from output_file
        for i in range(1, 21):  # Loop from 1 to 20 to fill rows 2 to 21 in output_file_new
            # Get the values from columns B, N, M, and O in output_file (B153 to B172, N153 to N172, M153 to M172, O153 to O172)
            value_af = ws[f'B{152+i}'].value
            value_ag = ws[f'N{152+i}'].value
            value_ah = ws[f'M{152+i}'].value
            value_ai = ws[f'O{152+i}'].value

            # Calculate the results to be written in output_file_new
            result_af = (value_af / divisor_7) / 1000 if divisor_7 else None
            result_ag = (value_ag / divisor_7) / 1000 if divisor_7 else None
            result_ah = (value_ah / divisor_7) / 1000 if divisor_7 else None
            result_ai = (value_ai / divisor_7) / 1000 if divisor_7 else None

            # Write the results to columns AF, AG, AH, and AI in output_file_new starting from row 2
            ws_new[f'AF{i+1}'] = result_af
            ws_new[f'AG{i+1}'] = result_ag
            ws_new[f'AH{i+1}'] = result_ah
            ws_new[f'AI{i+1}'] = result_ai

        # Copy values from output_file column B (from B321 to B340) to output_file_new column AJ (from AJ2 to AJ20)
        for i in range(321, 341):
            # Get the value from column B in output_file
            value = ws[f'B{i}'].value
            # Write the value to column AJ in output_file_new
            ws_new[f'AJ{i-319}'] = value
########################

        # Save the updated new workbook
        wb_new.save(output_file_new)
        print(f"Data copied and calculated successfully to {output_file_new}")
        return wb_new, ws_new
    except Exception as e:
        print(f"An error occurred while copying and calculating data: {e}")
        return None, None       

# Function to prepare data for database insertion
def data_preparation(df):
    sources = {
        'Soils': ['C', 'A', 'error', 'DISP', 'E'],
        'Sulphate': ['C', 'A', 'error', 'DISP', 'E'],
        'Diesel vehicles': ['C', 'A', 'error', 'DISP', 'E'],
        'Marine aerosol': ['C', 'A', 'error', 'DISP', 'E'],
        'Biomass burning': ['C', 'A', 'error', 'DISP', 'E'],
        'Petrol vehicles': ['C', 'A', 'error', 'DISP', 'E'],
        'Construction': ['C', 'A', 'error', 'DISP', 'E']
    }
    dfs = {}
    for source, measurements in sources.items():
        columns = ['Species'] + [f"{source}_{m}" for m in measurements]
        dfs[source] = df[columns].copy()
        new_columns = ['Species', "Concentration", "Average", "Error", "Dispersion", "Exceedance"]
        dfs[source].columns = new_columns

    df1 = dfs['Soils']
    df2 = dfs['Sulphate']
    df3 = dfs['Diesel vehicles']
    df4 = dfs['Marine aerosol']
    df5 = dfs['Biomass burning']
    df6 = dfs['Petrol vehicles']
    df7 = dfs['Construction']
    dfs = [df1, df2, df3, df4, df5, df6, df7] 
    return dfs
   
   
# Function to drop the existing database file
def create_database_p3(db_filename2):
    if os.path.exists(db_filename2):
        os.remove(db_filename2)
        print(f"Database file {db_filename2} has been deleted.")
    else:
        print(f"Database file {db_filename2} does not exist.")
    
    # Connect to the SQLite database
    conn = sqlite3.connect(db_filename2)
    cursor = conn.cursor()

    # Create Species table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS Species (
        SpeciesID INTEGER PRIMARY KEY AUTOINCREMENT,
        Species TEXT NOT NULL
    )
    ''')

    # Create Pollution Sourece table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS PollutionSources (
        SourceID TEXT PRIMARY KEY,
        Name TEXT NOT NULL
    )
    ''')
    # Insert pollution sources
    sources = [
        ('P1', 'Soils/RoadDust'),
        ('P2', 'Sulphate/MarineDiesel'),
        ('P3', 'DieselVehicles'),
        ('P4', 'MarineAerosol'),
        ('P5', 'BiomassBurning'),
        ('P6', 'PetrolVehicles'),
        ('P7', 'Construction')
    ]
    cursor.executemany('INSERT OR IGNORE INTO PollutionSources (SourceID, Name) VALUES (?, ?)', sources)

    # Create Samples table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS Samples (
        SampleID INTEGER PRIMARY KEY AUTOINCREMENT,
        Date TEXT NOT NULL,
        Time TEXT NOT NULL,
        Location TEXT NOT NULL
    )
    ''')

    # Create Measurement table
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS Measurement (
        MeasurementID INTEGER PRIMARY KEY AUTOINCREMENT,
        SampleID INTEGER,
        SpeciesID INTEGER,
        SourceID TEXT,
        Concentration REAL,
        Average REAL,
        Error REAL,
        Dispersion REAL,
        Exceedance REAL,
        FOREIGN KEY (SampleID) REFERENCES Samples (SampleID),
        FOREIGN KEY (SpeciesID) REFERENCES Species (SpeciesID),
        FOREIGN KEY (SourceID) REFERENCES PollutionSources (SourceID)
    )
    ''')

    # Commit the changes and close the connection
    conn.commit()
    conn.close()
    print(f"Database file {db_filename2} has been created and tables have been set up.")

def insert_Species_p3(db_filename2, df):
    # Connect to the SQLite database
    conn = sqlite3.connect(db_filename2)
    
    try:
        # Use the pandas to_sql method to insert data into the Species table
        df[['Species']].to_sql('Species', conn, if_exists='append', index=False)
        
        # Commit the changes
        conn.commit()
        print(f"Species data has been successfully inserted into the database {db_filename2}.")
    except Exception as e:
        print(f"An error occurred while inserting species data: {e}")
    finally:
        # Close the connection regardless of the operation outcome
        conn.close()
        

def insert_Measurement_p3(db_filename2, dfs, source_ids):
    # Connect to the SQLite database
    conn = sqlite3.connect(db_filename2)
    cursor = conn.cursor()

    # Obtain all SpeciesID
    cursor.execute('SELECT SpeciesID, Species FROM Species')
    species_dict = {name: id for id, name in cursor.fetchall()}

    # Define insert data function
    def insert_measurement(df, source_id, conn):
        inserted_count = 0
        try:
            cursor = conn.cursor()
            # Insert data
            for index, row in df.iterrows():
                # Obtain SpeciesID
                species_id = species_dict.get(row['Species'])
                if species_id is None:
                    print(f"Species {row['Species']} not found in Species table.")
                    continue

                # Insert Measurement data
                cursor.execute('''
                INSERT INTO Measurement (SampleID, SpeciesID, SourceID, Concentration, Average, Error, Dispersion, Exceedance)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (1, species_id, source_id,
                      row['Concentration'], row['Average'], row['Error'], row['Dispersion'], row['Exceedance']))
                inserted_count += 1

            # Confirm
            conn.commit()
            print(f"Successfully inserted {inserted_count} records into Measurement table for source {source_id}.")
        except sqlite3.Error as e:
            print("SQLite error: ", e)
            conn.rollback()
        except Exception as e:
            print("General error: ", e)
            conn.rollback()

    # Insert data
    for source, df in dfs.items():
        source_id = source_ids[source]
        insert_measurement(df, source_id, conn)

    # Close connection
    conn.close()
    print(f"Data insertion into {db_filename2} is complete.")

# Define function to plot pollution contribution
def plot_Pollution_Contribution(df_pollution_contribution):
    species_ids = list(range(1, 21))
    species_names = {
        1: 'H', 2: 'BC', 3: 'Na', 4: 'Mg',
        5: 'Al', 6: 'Si', 7: 'S', 8: 'Cl',
        9: 'K', 10: 'Ca', 11: 'Ti', 12: 'V',
        13: 'Mn', 14: 'Fe', 15: 'Co', 16: 'Ni',
        17: 'Cu', 18: 'Zn', 19: 'As', 20: 'Ba'
    }

    sources = ['P1', 'P2', 'P3', 'P4', 'P5', 'P6', 'P7']
    source_titles = [
        'Soil/Road Dust', 'Sulphate/Marine Diesel', 'Diesel Vehicles',
        'Marine Aerosol', 'Biomass Burning', 'Petrol Vehicles', 'Construction'
    ]

    fig, axes = plt.subplots(len(sources), 1, figsize=(8, 10), sharex=False)
    x = np.arange(len(species_ids))  # Label locations
    width = 0.75  # Width of the bars
    first_plot = True
    legend_handles = []

    for i, (source, title) in enumerate(zip(sources, source_titles)):
        ax = axes[i]
        source_data = df_pollution_contribution[df_pollution_contribution['SourceID'] == source]

        # Extract data for plotting
        concentration = source_data['Concentration'].values
        average = source_data['Average'].values
        error = source_data['Error'].values
        dispersion = source_data['Dispersion'].values
        exceedance = source_data['Exceedance'].values

        # Logarithmic y-axis for concentration
        ax.set_yscale('log')
        ax.set_ylim([1e-4, 1])
        ax.yaxis.set_major_formatter(FuncFormatter(lambda y, _: f'$10^{{{int(np.log10(y))}}}$'))

        # Right y-axis for Exceedance
        ax2 = ax.twinx()
        ax2.set_ylim([0, 100])
        ax2.set_yticks(np.arange(0, 101, 20))

        # Background vertical lines
        for j in x:
            ax.vlines(j, 1e-4, 1, colors='gray', linestyles='dashed', lw=0.5)

        # Plot data
        bars = ax.bar(x, concentration, width, color='lightblue', edgecolor='black')
        line = [ax.plot([x[j], x[j]], [error[j], dispersion[j]], color='black', lw=1) for j in range(len(x))]
        avg_line = ax.plot(x, average, color='red', marker='o', markersize=6, linestyle='', markerfacecolor='none')
        exceedance_line = ax2.plot(x, exceedance, color='black', marker='o', markersize=6, linestyle='')

        if first_plot:
            legend_handles.append(bars[0])  # Concentration
            legend_handles.append(avg_line[0])  # Average
            legend_handles.append(exceedance_line[0])  # Exceedance
            legend_handles.append(line[0][0])  # Dispersion
            first_plot = False

        # Set title and labels
        ax.text(1, 0.8, title, transform=ax.transAxes, ha='right', va='bottom', fontsize=10, weight='bold')
        ax.set_xticks(x)
        ax.set_xticklabels([species_names[sid] for sid in species_ids], rotation=0)

        # Add legend
        fig.legend(legend_handles, ['Concentration', 'Average', 'Exceedance', 'Maximum and Minimum DISP values'],
                loc='upper center', bbox_to_anchor=(0.5, -0.05), ncol=4)
        plt.tight_layout(pad=1)
        return fig




